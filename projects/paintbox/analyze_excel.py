#!/usr/bin/env python3
"""
Excel Expert Analysis Tool for BART 3.20 Workbook
Built with Anthropic's Claude for extracting and mapping 14,000+ formulas
"""

import json
import re
from pathlib import Path
from collections import defaultdict
from typing import Dict, List
import argparse

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Installing required packages...")
    import subprocess

    subprocess.check_call(["pip", "install", "openpyxl"])
    import openpyxl
    from openpyxl.utils import get_column_letter


class ExcelAnalyzer:
    """Analyzes Excel workbook and extracts all formulas with their dependencies"""

    def __init__(self, excel_path: str):
        self.excel_path = Path(excel_path)
        self.workbook = None
        self.formulas = defaultdict(list)
        self.named_ranges = {}
        self.cell_dependencies = defaultdict(set)
        self.formula_count = 0
        self.sheets_info = {}

    def load_workbook(self):
        """Load the Excel workbook"""
        print(f"Loading workbook: {self.excel_path}")
        self.workbook = openpyxl.load_workbook(self.excel_path, data_only=False)
        print(f"Loaded {len(self.workbook.sheetnames)} sheets")

    def analyze_named_ranges(self):
        """Extract all named ranges from the workbook"""
        print("\nAnalyzing named ranges...")
        try:
            # Handle different openpyxl versions
            if hasattr(self.workbook.defined_names, "definedName"):
                defined_names = self.workbook.defined_names.definedName
            elif hasattr(self.workbook, "defined_names"):
                defined_names = self.workbook.defined_names
            else:
                print("No defined names found")
                return

            for defined_name in defined_names:
                name = defined_name.name
                try:
                    destinations = (
                        list(defined_name.destinations)
                        if hasattr(defined_name, "destinations")
                        else []
                    )
                    self.named_ranges[name] = {
                        "name": name,
                        "scope": getattr(defined_name, "scope", None),
                        "value": defined_name.value if hasattr(defined_name, "value") else None,
                        "destinations": [
                            {"sheet": sheet, "range": range_str}
                            for sheet, range_str in destinations
                        ]
                        if destinations
                        else [],
                    }
                except Exception as e:
                    print(f"  Warning: Could not parse named range '{name}': {e}")

        except Exception as e:
            print(f"  Warning: Could not analyze named ranges: {e}")

        print(f"Found {len(self.named_ranges)} named ranges")

    def extract_cell_references(self, formula: str) -> List[str]:
        """Extract all cell references from a formula"""
        # Pattern to match cell references (e.g., A1, $A$1, Sheet1!A1)
        pattern = r"(?:\'[^\']+\'|[A-Za-z_]\w*)!?\$?[A-Z]+\$?\d+"
        references = re.findall(pattern, formula)
        return list(set(references))

    def categorize_formula(self, formula: str) -> str:
        """Categorize formula by its primary function"""
        formula_upper = formula.upper()

        # Financial functions
        if any(func in formula_upper for func in ["PMT", "PV", "FV", "RATE", "NPV", "IRR"]):
            return "Financial"

        # Lookup functions
        if any(
            func in formula_upper for func in ["VLOOKUP", "HLOOKUP", "INDEX", "MATCH", "XLOOKUP"]
        ):
            return "Lookup"

        # Statistical functions
        if any(func in formula_upper for func in ["AVERAGE", "STDEV", "VAR", "MEDIAN", "MODE"]):
            return "Statistical"

        # Math functions
        if any(func in formula_upper for func in ["SUM", "PRODUCT", "SQRT", "POWER", "LOG"]):
            return "Math"

        # Logical functions
        if any(func in formula_upper for func in ["IF", "AND", "OR", "NOT", "XOR"]):
            return "Logical"

        # Text functions
        if any(func in formula_upper for func in ["CONCATENATE", "LEFT", "RIGHT", "MID", "LEN"]):
            return "Text"

        # Date/Time functions
        if any(func in formula_upper for func in ["DATE", "TIME", "NOW", "TODAY", "YEAR", "MONTH"]):
            return "DateTime"

        # Basic arithmetic
        if re.match(r"^[=\+\-\*/\(\)\d\s\$A-Z]+$", formula_upper):
            return "Arithmetic"

        return "Other"

    def analyze_sheet(self, sheet_name: str):
        """Analyze all formulas in a specific sheet"""
        sheet = self.workbook[sheet_name]
        sheet_formulas = []

        print(f"\nAnalyzing sheet: {sheet_name}")
        print(f"  Dimensions: {sheet.max_row} rows x {sheet.max_column} columns")

        # Store sheet info
        self.sheets_info[sheet_name] = {
            "max_row": sheet.max_row,
            "max_column": sheet.max_column,
            "cell_count": sheet.max_row * sheet.max_column,
        }

        # Iterate through all cells
        for row in range(1, sheet.max_row + 1):
            for col in range(1, sheet.max_column + 1):
                cell = sheet.cell(row=row, column=col)

                if cell.value and isinstance(cell.value, str) and cell.value.startswith("="):
                    cell_ref = f"{get_column_letter(col)}{row}"
                    formula = cell.value

                    # Extract dependencies
                    dependencies = self.extract_cell_references(formula)
                    category = self.categorize_formula(formula)

                    formula_info = {
                        "sheet": sheet_name,
                        "cell": cell_ref,
                        "formula": formula,
                        "category": category,
                        "dependencies": dependencies,
                        "row": row,
                        "column": col,
                    }

                    sheet_formulas.append(formula_info)
                    self.formula_count += 1

                    # Track dependencies
                    for dep in dependencies:
                        self.cell_dependencies[dep].add(f"{sheet_name}!{cell_ref}")

        self.formulas[sheet_name] = sheet_formulas
        print(f"  Found {len(sheet_formulas)} formulas")

    def analyze_workbook(self):
        """Analyze the entire workbook"""
        print("\n" + "=" * 60)
        print("EXCEL FORMULA ANALYSIS")
        print("=" * 60)

        self.load_workbook()
        self.analyze_named_ranges()

        # Analyze each sheet
        for sheet_name in self.workbook.sheetnames:
            self.analyze_sheet(sheet_name)

        print(f"\nTotal formulas found: {self.formula_count}")

    def generate_category_summary(self) -> Dict[str, int]:
        """Generate summary of formulas by category"""
        category_count = defaultdict(int)

        for sheet_formulas in self.formulas.values():
            for formula in sheet_formulas:
                category_count[formula["category"]] += 1

        return dict(category_count)

    def find_complex_formulas(self, min_dependencies: int = 5) -> List[Dict]:
        """Find formulas with many dependencies"""
        complex_formulas = []

        for sheet_formulas in self.formulas.values():
            for formula in sheet_formulas:
                if len(formula["dependencies"]) >= min_dependencies:
                    complex_formulas.append(formula)

        return sorted(complex_formulas, key=lambda x: len(x["dependencies"]), reverse=True)

    def export_to_json(self, output_path: str):
        """Export analysis results to JSON"""
        output_path = Path(output_path)

        # Prepare data for export
        export_data = {
            "metadata": {
                "excel_file": str(self.excel_path),
                "total_formulas": self.formula_count,
                "sheet_count": len(self.workbook.sheetnames),
                "sheets_info": self.sheets_info,
                "category_summary": self.generate_category_summary(),
            },
            "named_ranges": self.named_ranges,
            "formulas_by_sheet": self.formulas,
            "complex_formulas": self.find_complex_formulas(),
            "dependencies": {
                cell: list(deps) for cell, deps in self.cell_dependencies.items() if len(deps) > 0
            },
        }

        # Write JSON file
        with open(output_path, "w") as f:
            json.dump(export_data, f, indent=2)

        print(f"\nAnalysis exported to: {output_path}")

    def generate_typescript_models(self, output_path: str):
        """Generate TypeScript interfaces for the formula structure"""
        output_path = Path(output_path)

        typescript_content = """// Auto-generated TypeScript models from Excel analysis
// Generated by Excel Expert Analysis Tool

export interface ExcelFormula {
  sheet: string;
  cell: string;
  formula: string;
  category: FormulaCategory;
  dependencies: string[];
  row: number;
  column: number;
}

export type FormulaCategory =
  | 'Financial'
  | 'Lookup'
  | 'Statistical'
  | 'Math'
  | 'Logical'
  | 'Text'
  | 'DateTime'
  | 'Arithmetic'
  | 'Other';

export interface NamedRange {
  name: string;
  scope?: string;
  destinations: Array<{
    sheet: string;
    range: string;
  }>;
  value?: string;
}

export interface SheetInfo {
  max_row: number;
  max_column: number;
  cell_count: number;
}

export interface ExcelAnalysis {
  metadata: {
    excel_file: string;
    total_formulas: number;
    sheet_count: number;
    sheets_info: Record<string, SheetInfo>;
    category_summary: Record<FormulaCategory, number>;
  };
  named_ranges: Record<string, NamedRange>;
  formulas_by_sheet: Record<string, ExcelFormula[]>;
  complex_formulas: ExcelFormula[];
  dependencies: Record<string, string[]>;
}
"""

        with open(output_path, "w") as f:
            f.write(typescript_content)

        print(f"TypeScript models generated: {output_path}")

    def generate_summary_report(self) -> str:
        """Generate a summary report of the analysis"""
        report_lines = [
            "\n" + "=" * 60,
            "EXCEL ANALYSIS SUMMARY REPORT",
            "=" * 60,
            f"\nWorkbook: {self.excel_path.name}",
            f"Total Sheets: {len(self.workbook.sheetnames)}",
            f"Total Formulas: {self.formula_count:,}",
            "\nFormulas by Category:",
        ]

        category_summary = self.generate_category_summary()
        for category, count in sorted(category_summary.items(), key=lambda x: x[1], reverse=True):
            percentage = (count / self.formula_count) * 100 if self.formula_count > 0 else 0
            report_lines.append(f"  {category:<15} {count:>6,} ({percentage:>5.1f}%)")

        report_lines.extend(
            [
                f"\nNamed Ranges: {len(self.named_ranges)}",
                "\nTop 5 Most Complex Formulas (by dependency count):",
            ]
        )

        complex_formulas = self.find_complex_formulas()[:5]
        for i, formula in enumerate(complex_formulas, 1):
            report_lines.append(
                f"  {i}. {formula['sheet']}!{formula['cell']} - "
                f"{len(formula['dependencies'])} dependencies"
            )

        report_lines.extend(["\nSheets Overview:"])

        for sheet_name, info in self.sheets_info.items():
            formula_count = len(self.formulas.get(sheet_name, []))
            report_lines.append(
                f"  {sheet_name:<30} {info['max_row']:>6} rows x {info['max_column']:>4} cols, "
                f"{formula_count:>6,} formulas"
            )

        report = "\n".join(report_lines)
        print(report)
        return report


def main():
    parser = argparse.ArgumentParser(
        description="Excel Expert Analysis Tool - Extract and analyze Excel formulas"
    )
    parser.add_argument(
        "--excel",
        "-e",
        default="/Users/patricksmith/candlefish-ai/projects/bart-estimator-pwa/data/bart3.20.xlsx",
        help="Path to Excel file",
    )
    parser.add_argument(
        "--output", "-o", default="excel_analysis.json", help="Output JSON file path"
    )
    parser.add_argument(
        "--typescript",
        "-t",
        default="excel_models.ts",
        help="Output TypeScript models file",
    )

    args = parser.parse_args()

    # Create analyzer instance
    analyzer = ExcelAnalyzer(args.excel)

    # Run analysis
    analyzer.analyze_workbook()

    # Generate outputs
    analyzer.export_to_json(args.output)
    analyzer.generate_typescript_models(args.typescript)

    # Print summary report
    analyzer.generate_summary_report()

    print(f"\n✅ Analysis complete! Check {args.output} for full results.")


if __name__ == "__main__":
    main()
